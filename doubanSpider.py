import time
import urllib

import requests
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import Workbook

# Some User Agents
headers = [{'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) '
                          'Gecko/20091201 Firefox/3.5.6'},
           {'User-Agent': 'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) '
                          'Chrome/17.0.963.12 Safari/535.11'},
           {'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}]


def book_spider(book_tag):
    page_num = 0
    book_list = []
    try_times = 0

    while page_num < 2:
        # url='http://www.douban.com/tag/%E5%B0%8F%E8%AF%B4/book?start=0' # For Test
        url = 'http://www.douban.com/tag/' + urllib.parse.quote(book_tag) + '/book?start=' + str(page_num * 15)
        time.sleep(np.random.rand() * 5)

        try:
            html = requests.get(url, headers=headers[1])
        except Exception as e:
            print("[info] some error occur")
            continue

        soup = BeautifulSoup(html.text, 'lxml')
        list_soup = soup.find('div', class_='mod book-list')

        if list_soup is None and try_times < 200:
            continue
        elif list_soup is None or len(list_soup) <= 1:
            break  # Break when no information got after 200 times requesting

        for book_info in list_soup.find_all('dd'):
            # 提取书名
            title = book_info.find('a', class_='title').string.strip()
            desc = book_info.find('div', class_='desc').string.strip()
            desc_list = desc.split('/')
            # 提取书的链接
            book_url = book_info.find('a', class_='title').get('href')

            # 提取书的作者/译者
            # noinspection PyBroadException
            try:
                author_info = '/'.join(desc_list[0:-3])
            except:
                author_info = '暂无'

            # 提取书的出版信息
            # noinspection PyBroadException
            try:
                pub_info = '/'.join(desc_list[-3:])
            except:
                pub_info = '暂无'

            # 提取书的评分
            # noinspection PyBroadException
            try:
                rating = float(book_info.find('span', class_='rating_nums').string.strip())
            except:
                rating = 0.0

            book_list.append([title, rating, author_info, pub_info])
            try_times = 0

        page_num += 1
        print('Downloading Information From Page %d' % page_num)
    return book_list


def do_spider(book_tag_lists):
    book_lists = []
    for book_tag in book_tag_lists:
        book_list = book_spider(book_tag)
        book_list = sorted(book_list, key=lambda x: x[1], reverse=True)
        book_lists.append(book_list)
    return book_lists


def lists_to_excel(book_lists, book_tag_lists):
    wb = Workbook()  # optimized_write = True是一个优化的快速写入方法
    ws = []
    for index in range(len(book_tag_lists)):
        ws.append(wb.create_sheet(title=book_tag_lists[index]))
    for index in range(len(book_tag_lists)):
        ws[index].append(['序号', '书名', '评分', '作者/译者', '出版社'])
        count = 1
        for book_list in book_lists[index]:
            ws[index].append([count, book_list[0], book_list[1], book_list[2], book_list[3]])
            count += 1
    save_path = 'book_list'
    for i in range(len(book_tag_lists)):
        save_path += ('-' + book_tag_lists[i])
    save_path += '.xlsx'
    wb.save(save_path)


if __name__ == '__main__':
    book_tags = ['小说', '心理']
    books = do_spider(book_tags)
    lists_to_excel(books, book_tags)
